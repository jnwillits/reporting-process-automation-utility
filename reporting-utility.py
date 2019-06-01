# !/usr/bin/env python
"""
Jeff's Reporting Process Automation Utility - Version 1.0"
This is an example program that demonstrates how data can be automatically collected from multiple spreadsheets. The
program processes the data and generates report files in Excel and PDF formats. The Python source code for this utility
is available in my GitHub repository.   Jeffrey Neil Willits     jnwillits.com
"""

import PySimpleGUI as sg
import json
import os
import sys
import openpyxl
from datetime import datetime
from dateutil.parser import *
from win32com import client
import win32api

sg.ChangeLookAndFeel('Dark')
sg.SetOptions(icon='aplanner_icon.ico', element_padding=(6, 6), font=('verdana', 9), text_color='#32CD32',
              background_color='#1E1E1E', text_element_background_color='#1E1E1E', button_color=('#FFFFFF', '#2F2F2F'))
menu_def = [['Setup', ['Honolulu', 'Seattle', 'Denver']],
            ['Help', 'About...']]

layout = [
    [sg.Menu(menu_def, tearoff=False, pad=(20, 1))],
    [sg.T('')],
    [sg.T('Honolulu file path:', text_color='#FFFFFF', font=('verdana', 8))],
    [sg.T('', size=(200, 1), key='_HONOLULU_FILE_PATH_', font=('verdana', 8)), ],
    [sg.T('')],
    [sg.T('')],
    [sg.T('Seattle file path:', text_color='#FFFFFF', font=('verdana', 8))],
    [sg.T('', size=(200, 1), key='_SEATTLE_FILE_PATH_', font=('verdana', 8)), ],
    [sg.T('')],
    [sg.T('')],
    [sg.T('Denver file path:', text_color='#FFFFFF', font=('verdana', 8))],
    [sg.T('', size=(200, 1), key='_DENVER_FILE_PATH_', font=('verdana', 8)), ],
    [sg.T('')],
    [sg.T('')],
    [sg.Button('', visible=False, size=(12, 1), ), ],
    [sg.Button('Make Report', visible=True, size=(12, 1), )],
    [sg.Button('Cancel', visible=True, size=(12, 1), ), ]]


def define_file():
    if len(sys.argv) == 1:
        event_define_file, (file_path,) = sg.Window('My Script').Layout([[sg.Text('Document to open')],
                                                                         [sg.In(size=(50, 10)), sg.FileBrowse()],
                                                                         [sg.CloseButton('Open'),
                                                                          sg.CloseButton('Cancel')]]).Read()
    else:
        file_path = sys.argv[1]
    if not file_path:
        sg.Popup("Cancel", "No file path was supplied.", background_color='#4a646c', text_color='#ffffff')
        raise SystemExit("Cancelling - no file path was supplied.")
    return file_path


def read_files():
    if os.path.isfile('filepath.json'):
        with open('filepath.json') as f_obj:
            return json.load(f_obj)
    else:
        file_path_honolulu = define_file()
        file_path_seattle = define_file()
        file_path_denver = define_file()
        return {'honolulu': file_path_honolulu, 'seattle': file_path_seattle, 'denver': file_path_denver}


def write_data(file_path_dict_pass):
    with open('filepath.json', 'w') as f_obj:
        json.dump(file_path_dict_pass, f_obj)


def read_col(ws, start_row, start_col, name_col):
    col_list = []
    last_row = len(ws[name_col])
    for row in ws.iter_rows(min_row=start_row, min_col=start_col, max_row=last_row, max_col=start_col):
        for cell in row:
            if cell.value is not None:
                col_list.append(cell.value)
    return col_list


def populate_data_objects(f_str, sheet_name, office, start_row, start_col):
    """ Extract data from the spreadsheets. Then populate lists of tuples. The tuples hold data associated with
        each employee. """
    wb = openpyxl.load_workbook(f_str)
    ws = wb[sheet_name]
    name_col_list = read_col(ws, start_row, start_col, name_col='B')
    revenue_col_list = read_col(ws, start_row, start_col + 1, name_col='B')
    hours_col_list = read_col(ws, start_row, start_col + 2, name_col='B')
    wb.close()
    location_data = []
    for i in range(0, len(name_col_list)):
        if revenue_col_list[i] != 0 and hours_col_list[i] != 0:
            productivity = revenue_col_list[i] / hours_col_list[i]
        else:
            productivity = 0
        location_data.append((office, name_col_list[i], revenue_col_list[i], hours_col_list[i], productivity))
    return location_data


def core_tasks(ws1_file, ws2_file, ws3_file):
    """ This is the main operational function. """
    agent_max_list = []
    team_average_list = []

    # HONOLULU
    honolulu_list = populate_data_objects(ws1_file, 'Sheet1', office='HONOLULU', start_row=6, start_col=2)
    honolulu_list.sort(key=lambda item: item[4], reverse=True)
    agent_max_list.append(honolulu_list[0][4])
    productivity_list = []
    for i in range(0, len(honolulu_list)):
        productivity_list.append(honolulu_list[i][4])
    team_average_list.append(sum(productivity_list) / len(honolulu_list))

    # SEATTLE
    seattle_list = populate_data_objects(ws2_file, 'Sheet1', office='SEATTLE', start_row=6, start_col=2)
    seattle_list.sort(key=lambda item: item[4], reverse=True)
    agent_max_list.append(seattle_list[0][4])
    productivity_list = []
    for i in range(0, len(seattle_list)):
        productivity_list.append(seattle_list[i][4])
    team_average_list.append(sum(productivity_list) / len(seattle_list))

    # DENVER
    denver_list = populate_data_objects(ws3_file, 'Sheet1', office='DENVER', start_row=6, start_col=2)
    denver_list.sort(key=lambda item: item[4], reverse=True)
    agent_max_list.append(denver_list[0][4])
    productivity_list = []
    for i in range(0, len(denver_list)):
        productivity_list.append(denver_list[i][4])
    team_average_list.append(sum(productivity_list) / len(denver_list))

    combined_list = honolulu_list + seattle_list + denver_list

    f_str = 'daily-report.xlsx'
    wb = openpyxl.load_workbook(f_str)

    # Populate the Data sheet (sorted by revenue).
    ws = wb['Data']
    start_col = 2
    start_row = 7
    combined_list.sort(key=lambda item: item[2], reverse=True)
    for i in range(0, len(combined_list)):
        ws.cell(row=start_row + i, column=start_col).value = combined_list[i][0]
        ws.cell(row=start_row + i, column=start_col + 1).value = combined_list[i][1]
        ws.cell(row=start_row + i, column=start_col + 2).value = combined_list[i][2]
        ws.cell(row=start_row + i, column=start_col + 3).value = combined_list[i][3]
        ws.cell(row=start_row + i, column=start_col + 4).value = combined_list[i][4]
    ws.delete_rows(84)

    # Populate the top agents list (sorted by productivity).
    ws = wb['Report']
    start_col = 2
    start_row = 9
    combined_list.sort(key=lambda item: item[4], reverse=True)
    for i in range(0, 18):
        ws.cell(row=start_row + i, column=start_col).value = combined_list[i][0]
        ws.cell(row=start_row + i, column=start_col + 1).value = combined_list[i][1]
        ws.cell(row=start_row + i, column=start_col + 2).value = combined_list[i][4]

    # Populate bar graph data cells.
    for i in range(0, 3):
        ws.cell(row=33, column=7 + i).value = agent_max_list[i]
    for i in range(0, 3):
        ws.cell(row=34, column=7 + i).value = team_average_list[i]

    # Save the consolidated data/report as an Excel file.
    date_time_str = datetime.now().strftime(' %Y-%m-%d %H %M %S')
    report_file_str = f_str[:-5] + date_time_str
    wb.save(report_file_str + '.xlsx')
    wb.close()

    # Save the report as a PDF file.
    path_str = 'c:/users/jeffw/documents/code/work/reporting utility/'
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = 0
    wb = excel.Workbooks.Open(path_str + report_file_str + '.xlsx')
    try:
        wb.SaveAs('c:\\users\\jeffw\\documents\\code\\work\\reporting utility\\' + report_file_str, FileFormat=57)
    except Exception as e:
        sg.Print("Failed to convert", str(e))
    finally:
        wb.Close()
        excel.Quit()

    sys.exit()


if __name__ == '__main__':
    file_path_dict = read_files()
    ws1_file_path = file_path_dict['honolulu']
    ws2_file_path = file_path_dict['seattle']
    ws3_file_path = file_path_dict['denver']
    path_list = [ws1_file_path, ws2_file_path, ws3_file_path]
    path_length = len(max(path_list))
    window = sg.Window(" Jeff's Reporting Utility", size=(path_length + 800, 600), default_element_size=(30, 1),
                       grab_anywhere=False, background_color='#1E1E1E', auto_size_text=False,
                       auto_size_buttons=False).Layout(layout).Finalize()
    window.Element('_HONOLULU_FILE_PATH_').Update(ws1_file_path)
    window.Element('_SEATTLE_FILE_PATH_').Update(ws2_file_path)
    window.Element('_DENVER_FILE_PATH_').Update(ws3_file_path)

while True:
    event, values = window.Read(timeout=10)
    if event is None or event == 'Exit':
        break
    else:
        if event == 'About...':
            sg.Popup("Jeff's Reporting Process Automation Utility - Version 1.0",
                     'This is an example program that demonstrates how data can be \n'
                     'automatically collected from multiple spreadsheets. The program \n'
                     'processes the data and generates report files in Excel and PDF \n'
                     'formats. The Python source code for this utility is available \n'
                     'in my GitHub repository.\n\n'
                     'Jeffrey Neil Willits', 'jnwillits.com\n', no_titlebar=True, keep_on_top=True,
                     grab_anywhere=True, background_color='#4a646c', text_color='#ffffff')
        elif event == 'Honolulu':
            ws1_file_path = define_file()
            window.Element('_HONOLULU_FILE_PATH_').Update(ws1_file_path)
            file_path_dict['honolulu'] = ws1_file_path
            write_data(file_path_dict)
        elif event == 'Seattle':
            ws2_file_path = define_file()
            window.Element('_SEATTLE_FILE_PATH_').Update(ws2_file_path)
            file_path_dict['seattle'] = ws2_file_path
            write_data(file_path_dict)
        elif event == 'Denver':
            ws3_file_path = define_file()
            window.Element('_DENVER_FILE_PATH_').Update(ws3_file_path)
            file_path_dict['denver'] = ws3_file_path
            write_data(file_path_dict)
        elif event == 'Make Report':
            core_tasks(ws1_file_path, ws2_file_path, ws3_file_path)
        elif event == 'Cancel':
            window.Close()
            sys.exit()

window.Close()
sys.exit()
